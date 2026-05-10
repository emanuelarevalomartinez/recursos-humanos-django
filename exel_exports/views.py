import os
import openpyxl
import openpyxl.workbook 
import json
from django.shortcuts import render
from django.http import JsonResponse
from django.views import View
from django.conf import settings
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles.borders import Border, Side
from graphics.models import History
from datetime import datetime
from .models import Data, Ae2, Ae3
# Create your views here.
class CreateAe2(View):
    
    template = 'document/Ae2.html'

    def get(self,request):
        #consultas a la base de datos
        data = Data.objects
        in2 = data.filter(tipo_de_contrato = 'Indeterminado').count()
        in3 =  data.all().count()
        in4 = data.filter(categoria_di =['Profesor Auxiliar','Profesor Titular']).filter(tipo_de_contrato = 'Determinado').count()
        in5 = data.filter(categoria_di =['Profesor Auxiliar','Profesor Titular']).filter(tipo_de_contrato = ['Determinado','Indeterminado']).count()
        in6 = data.exclude(subcategoria = ['DOCENTES  PRINCIPAL','Docente','Docente Pincipal']).count()
        in8 = data.exclude(subcategoria = ['DOCENTES  PRINCIPAL','Docente','Docente Pincipal']).filter(tipo_de_contrato='Provisional').count()
        in9 = data.exclude(subcategoria = ['DOCENTES  PRINCIPAL','Docente','Docente Pincipal']).filter(tipo_de_contrato='A Prueba').count()
        in10 = data.filter(tipo_de_contrato = 'Determinado').count()
        in12 = data.exclude(subcategoria = ['DOCENTES  PRINCIPAL','Docente','Docente Pincipal']).filter(causa_alta = '13').count()
        in13 = data.exclude(subcategoria = ['DOCENTES  PRINCIPAL','Docente','Docente Pincipal']).filter(subcategoria = ['Otros Tecnicos','OTROS TECNICOS']).count()
        in21 = data.filter(tipo_de_contrato = 'Adiestramiento').count()
        in24 = data.filter(salario_basico = 3610).count()
       
        predata = {
           'indicador1' : 0,
           'indicador2' : in2,
           'indicador3' : in3,
           'indicador4' : in4,
           'indicador5' : in5,
           'indicador6' : in6,
           'indicador7' : in6,
           'indicador8' : in8,
           'indicador9' :in9,
           'indicador10' : in10,
           'indicador11' : 0,
           'indicador12' : in12,
           'indicador13' : in13,
           'indicador14' : 0,
           'indicador15' : 0,
           'indicador16' : 0,
           'indicador17' : 0,
           'indicador18' : 0,
           'indicador19' : in21,
           'indicador20' : 0,
           'indicador21' : in21,
           'indicador22' : in24,
           'indicador23' : in24,
        }
        return render(request,self.template,{
           'predata': predata,
        })
    
    def post(self,request):
        try:
         #datos input
          data = json.loads(request.body.decode('utf-8'))
          total_p = data.get('total_plantilla_aprobada')
          total_c = data.get('total_plantilla_cubierta')
          total_g = data.get('total_general_de_contrato')
          total_c_p_det = data.get('total_contrato_p_det')
          total_contratopc = data.get('total_contrato_p_comp')
          total_contratonod = data.get('total_contrato_no_d')
          contrato_no_doc_r = data.get('contrato_no_doc_resp')
          contr_no_rs = data.get('contrato_no_doc_resp_s')
          periodo_p = data.get('periodo_prueba')
          seren = data.get('seren_aux')
          lab_a = data.get('labores_a')
          jubilados = data.get('jubilados')
          otros = data.get('otros')
          contr_no_d =data.get('contratos_no_d')
          contr_no_d_e = data.get('contratos_no_d_e')
          eject_obr = data.get('eject_obra')
          rese_cint = data.get('reserva_cientif')
          recien_g = data.get('recien_graduados')
          reserva = data.get('reserva_d_p')
          tecn_a_p = data.get('tecnico_m_prepa')
          estu_c_d =data.get('estu_contrd_d')
          estu_a = data.get('estu_cont_a')
          estu_co_n = data.get('estu_cont_n')
          confecc = data.get('confeccionado')
          c_carg = data.get('c_cargo')
          revis = data.get('revisado')
          r_carg = data.get('r_cargo')
   
      

          workbook = openpyxl.Workbook()
          worksheet = workbook.active

          #datos de entrada
          worksheet['B11'] = total_p
          worksheet['B12'] = total_c 
          worksheet['B13'] = total_g
          worksheet['B14'] = total_c_p_det
          worksheet['B15'] = total_contratopc
          worksheet['B16'] = total_contratonod
          worksheet['B17'] = contrato_no_doc_r
          worksheet['B18'] = contr_no_rs
          worksheet['B19'] = periodo_p
          worksheet['B20'] = seren
          worksheet['B21'] = lab_a
          worksheet['B22'] = jubilados
          worksheet['B23'] = otros
          worksheet['B24'] = contr_no_d
          worksheet['B25'] = contr_no_d_e
          worksheet['B26'] = eject_obr
          worksheet['B27'] = rese_cint
          worksheet['B28'] = recien_g
          worksheet['B29'] = reserva
          worksheet['B30'] = tecn_a_p
          worksheet['B31'] = estu_c_d
          worksheet['B32'] = estu_a
          worksheet['B33'] = estu_co_n 
          
          worksheet['F11'] = total_p
          worksheet['F12'] = total_c 
          worksheet['F13'] = total_g
          worksheet['F14'] = total_c_p_det
          worksheet['F15'] = total_contratopc
          worksheet['F16'] = total_contratonod
          worksheet['F17'] = contrato_no_doc_r
          worksheet['F18'] = contr_no_rs
          worksheet['F19'] = periodo_p
          worksheet['F20'] = seren
          worksheet['F21'] = lab_a
          worksheet['F22'] = jubilados
          worksheet['F23'] = otros
          worksheet['F24'] = contr_no_d
          worksheet['F25'] = contr_no_d_e
          worksheet['F26'] = eject_obr
          worksheet['F27'] = rese_cint
          worksheet['F28'] = recien_g
          worksheet['F29'] = reserva
          worksheet['F30'] = tecn_a_p
          worksheet['F31'] = estu_c_d
          worksheet['F32'] = estu_a
          worksheet['F33'] = estu_co_n 
          


          #estilos
          bold_font = Font(bold=True, name='Areal',size=10)
          areal = Font(name = 'Areal', size=10)
          center_alignment = Alignment(horizontal='center')
          worksheet.column_dimensions['A'].width = 80
          thick_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))


          columns_to_adjust = {
             'B': 6,  
             'C': 6,  
             'D': 6,
             'E': 6,
             'F': 7,
            }
          
          for letter, width in columns_to_adjust.items():
              column = worksheet.column_dimensions[letter]
              column.width = width

          for row in worksheet['A10:F33']:
              for cell in row:
                  cell.border = thick_border

          #estandarizaciones
          worksheet['A3'] = 'MODELO 223.216 (I) AE-2'
          worksheet['A3'].font = bold_font
          worksheet['A5'] = 'INDICADORES MENSUALES DE LA ACTIVIDAD DE RECURSOS HUMANOS'
          worksheet['A5'].font = bold_font
          worksheet['A5'].alignment = center_alignment
          worksheet['A7'] = 'CENTRO INFORMANTE:'
          worksheet['A7'].font = bold_font
          worksheet['A9'] = f'MES: {datetime.now().month} {datetime.now().year}'
          worksheet['A9'].font = bold_font
          worksheet['C9'] = f'FECHA:{datetime.now().day}-{datetime.now().month}-{datetime.now().year}'
          worksheet['C9'].font = bold_font
          
          #tabla
          worksheet['A10'] = 'Indicadores'
          worksheet['A10'].font = bold_font
          worksheet['A10'].alignment = center_alignment
          worksheet['B10'] = 'I'
          worksheet['B10'].font = bold_font
          worksheet['B10'].alignment = center_alignment
          worksheet['C10'] = 'II'
          worksheet['C10'].font = bold_font
          worksheet['C10'].alignment = center_alignment
          worksheet['D10'] = 'III'
          worksheet['D10'].font = bold_font
          worksheet['D10'].alignment = center_alignment
          worksheet['E10'] = 'IV'
          worksheet['E10'].font = bold_font
          worksheet['E10'].alignment = center_alignment
          worksheet['F10'] = 'TOTAL'
          worksheet['F10'].font = bold_font
          worksheet['F10'].alignment = center_alignment
          
          indc = {
               'A11' : '1.- Total Plantilla Aprobada',
               'A12' : '2.- Total Plantilla Cubierta',
               'A13' : '3.- Total General de Contratos',
               'A14' : '4.- Total de Contratos de Profesores por tiempo determinado',
               'A15' : '5.- De ellos: a tiempo completo',
               'A16' : '6.- Total de Contratos No Docentes',
               'A17' : '7.- Contratos No Docentes con respaldo de plazas',
               'A18' : '8.- De ellos: por sustitución',
               'A19' : '9.- Período de Prueba',
                'A20' : '10.- Serenos y Auxiliares de Limpieza',
                'A21' : '11.- Labores Agrícolas',
                'A22' : '12.- Jubilados',
                'A23' : '13.- Otros',
                'A24' : '14.- Contratos No Docentes sin respaldo de plazas (15 a 19)',
                'A25' : '15.- De ellos: Serenos y Auxiliares de Limpieza',
                'A26' : '16.- Ejecución de Obra',
                'A27' : '17.- Reserva Científica en Preparación',
                'A28' : '18.- Recién Graduados en Preparación',
                'A29' : '19.- Reserva Dirección Provincial de Trabajo',
                'A30' : '20.- Técnicos Medios en Preparación',
                'A31' : '21 - Total de estudiantes de la Universidad de CD contratados por tiempo determinado',
                'A32' : '22 - Del total de estudiantes de CD contratados, cifras como Auxiliar Técnico de la Docencia',
                'A33' : '23 - Del total de estudiantes de CD contratados, cifras en cargos No Docentes'

            }

          for key, value in indc.items():
              worksheet[key] = value
              worksheet[key].font = areal

          
          worksheet['A38'] = 'Observaciones:'
          worksheet['A38'].font = bold_font
          worksheet['A40'] = f'Confeccionado por:    {confecc}                                     Revisado por:{revis}'
          worksheet['A41'] = f'                           {c_carg}                                                               {r_carg}'
          worksheet['A43'] = '                                                                        DrC. Raydel Montesino Perurena'
          worksheet['A44'] = '                                                                                Rector UCI'


          for row in worksheet['F11:F33']:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
         
          for row in worksheet['B13:E13']:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

          for row in worksheet['B16:E16']:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

          for row in worksheet['B17:E17']:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

          for row in worksheet['B24:E24']:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

          for row in worksheet['B31:F31']:
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')


          filename = 'Ae2.xlsx'
          filepath = os.path.join(settings.BASE_DIR, 'static', 'informes', filename)
    
          workbook.save(filepath)

          history = History.objects.create( type_document = 'Ae2')
          print(history.date)
          history.save()
          
         
          ae2 = Ae2.objects.create(
             total_platilla_aprobada = total_p,
             total_platilla_cubierta = total_c,
             total_general_contratos = total_g,
             total_contrato_P_det = total_c_p_det,
             total_contratos_P_comp = total_contratopc,
             total_contrato_no_d = total_contratonod,
             contrato_no_doc_resp = contrato_no_doc_r,
             contrato_no_doc_resp_s = contr_no_rs,
             periodo_prueba = periodo_p,
             seren_aux = seren,
             labores_a = lab_a,
             jubilados = jubilados,
             otros = otros,
             contratos_no_d = contr_no_d,
             contratos_no_d_e = contr_no_d_e,
             eject_obra = eject_obr,
             reserva_cientif = rese_cint,
             recien_graduados = recien_g,
             reserva_d_p =  reserva,
             tecnico_m_prepa = tecn_a_p,
             estu_contr_d = estu_c_d,
             estu_cont_a = estu_a,
             estu_cont_n = estu_co_n,
             confeccionado = confecc,
             c_cargo = c_carg,
             revisado = revis,
             r_cargo = r_carg
          )

          ae2.save()

          return JsonResponse({
              "ok": True,
              "message" :"file created"
          })
        except Exception as error:
            return JsonResponse({
                'ok': False,
                'message': str(error)
            })

   



class CreateAe3(View):

    template = 'document/Ae3.html'
    count  = 1
    count2 = 1
    def get(self,request):
         
         #consultas a la base de datos
        ae3 = Data.objects
        in1 = ae3.all().count()
        in2 = ae3.filter(subcategoria = ['DOCENTES  PRINCIPAL','Docente','Docente Pincipal']).count()
        in3 = ae3.filter(categoria = 'Cuadro Eject.').count()
        in4 = ae3.filter(categoria = 'investigativo').count()
        in5 = ae3.filter(subcategoria = ['Otros Tecnicos','OTROS TECNICOS']).count()
        in6 = ae3.filter(categoria = 'Tecnicos').count()
        in7 = ae3.filter(categoria_di =['Profesor Auxiliar','Profesor Titular']).filter(tipo_de_contrato = 'Indeterminado').count()
        in8 = ae3.filter(categoria = 'Asesores y Metodologos').count()
        in9 = ae3.filter(cargo = 'Investigativo').count()
        in12 = ae3.filter(cargo = 'Servicio').count()
        in13 = ae3.filter(categoria = 'Operarios').count()
        in15 = ae3.filter(categoria_di =['Profesor Auxiliar','Profesor Titular']).filter(tipo_de_contrato = 'A Tpo Parcial').count()
        
        in16 = ae3.all().filter(sexo = 'F').count()
        in17 = ae3.all().filter(menor_o_igual_35 = True).count()
        in18 = ae3.all().filter(sexo = 'F').filter(menor_o_igual_35 = True).count()
        in19 = ae3.all().filter(categoria_di = 'Profesor Titular').count()
        in20 = ae3.all().filter(categoria_di = 'Profesor Auxiliar').count()
        in21 = ae3.all().filter(categoria_di = 'Asistente').count()
        in22 = ae3.all().filter(categoria_di = 'Instructor').count()

        in23 = ae3.filter(subcategoria = ['DOCENTES  PRINCIPAL','Docente','Docente Pincipal']).filter(sexo = 'F').count()
        in24 = ae3.filter(subcategoria = ['DOCENTES  PRINCIPAL','Docente','Docente Pincipal']).filter(menor_o_igual_35 = True).count()
        in25 = ae3.filter(subcategoria = ['DOCENTES  PRINCIPAL','Docente','Docente Pincipal']).filter(sexo = 'F').filter(menor_o_igual_35 = True).count() 
        in26 = ae3.filter(subcategoria = ['DOCENTES  PRINCIPAL','Docente','Docente Pincipal']).filter(categoria_di = 'Profesor Titular').count()
        in27 = ae3.filter(subcategoria = ['DOCENTES  PRINCIPAL','Docente','Docente Pincipal']).filter(categoria_di = 'Profesor Auxiliar').count()
        in28 = ae3.filter(subcategoria = ['DOCENTES  PRINCIPAL','Docente','Docente Pincipal']).filter(categoria_di = 'Asistente').count()
        in29 = ae3.filter(subcategoria = ['DOCENTES  PRINCIPAL','Docente','Docente Pincipal']).filter(categoria_di = 'Instructor').count()

        in30 = ae3.filter(categoria = 'Tecnicos').filter(sexo = 'F').count()
        in31 = ae3.filter(categoria = 'Tecnicos').filter(menor_o_igual_35 = True).count()
        in32 = ae3.filter(categoria = 'Tecnicos').filter(sexo = 'F').filter(menor_o_igual_35 = True).count() 
        in33 = ae3.filter(categoria = 'Tecnicos').filter(categoria_di = 'Profesor Titular').count()
        in34 = ae3.filter(categoria = 'Tecnicos').filter(categoria_di = 'Profesor Auxiliar').count()
        in35 = ae3.filter(categoria = 'Tecnicos').filter(categoria_di = 'Asistente').count()
        in36 = ae3.filter(categoria = 'Tecnicos').filter(categoria_di = 'Instructor').count()

        predata = {
           'indicador1' : in1,
           'indicador2' : in2,
           'indicador3' : in3,
           'indicador4' : in4,
           'indicador5' : in5,
           'indicador6' : in6,
           'indicador7' : in7,
           'indicador8' : in8,
           'indicador9' : in9,
           'indicador10' : in5,
           'indicador11' : in3,
           'indicador12' : in12,
           'indicador13' : in13,
           'indicador14' : in1,
           'indicador15' : in15,
           'indicador16' : in16,
           'indicador17' : in17,
           'indicador18' : in18,
           'indicador19' : in19,
           'indicador20' : in20,
           'indicador21' : in21,
           'indicador22' : in22,
           'indicador23' : in23,
           'indicador24' : in24,
           'indicador25' : in25,
           'indicador26' : in26,
           'indicador27' : in27,
           'indicador28' : in28,
           'indicador29' : in29,
           'indicador30' : in30,
           'indicador31' : in31,
           'indicador32' : in32,
           'indicador33' : in33,
           'indicador34' : in34,
           'indicador35' : in35,
           'indicador36' : in36,
         }
        return render(request,self.template,{
           'predata': predata
        })

    def post(self, request):
        try:
            data = json.loads(request.body.decode('utf-8'))


            confe = data.get('confeccionado')
            aprob = data.get('aprobado')

            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            
            worksheet['C12'] = data.get('total_de_cuadros')
            worksheet['C13'] = data.get('cuadros_docente')
            worksheet['C14'] = data.get('cuadros_admin')
            worksheet['C15'] = data.get('cuadros_invest')
            worksheet['C16'] = data.get('otros_cuadros')
            worksheet['C17'] = data.get('total_tecnicos')
            worksheet['C18'] = data.get('prof_tiempo_compl')
            worksheet['C19'] = data.get('asesores_metodol')
            worksheet['C20'] = data.get('investigadores')
            worksheet['C21'] = data.get('otros_tecnicos')
            worksheet['C22'] = data.get('administrativos')
            worksheet['C23'] = data.get('servicio')
            worksheet['C24'] = data.get('operarios')
            worksheet['C25'] = data.get('total')
            worksheet['C26'] = data.get('profe_tiempo_pars')
            worksheet['D12'] = data.get('total_de_cuadros_fem')
            worksheet['E12'] = data.get('total_jovenes_t')
            worksheet['F12'] = data.get('total_jovenes_t_fem')
            worksheet['G12'] = data.get('profe_titular')
            worksheet['H12'] = data.get('profe_aux')
            worksheet['I12'] = data.get('asistente')
            worksheet['J12'] = data.get('instructor')

            worksheet['D13'] = data.get('cuadros_docente_fem')
            worksheet['E13'] = data.get('cuadros_docente_jovenes_t')      
            worksheet['F13'] = data.get('cuadros_docente_jovenes_t_fem') 
            worksheet['G13'] = data.get('cuadros_docente_profe_titular')
            worksheet['H13'] = data.get('cuadros_docente_profe_aux')
            worksheet['I13'] = data.get('cuadros_docente_asistente')
            worksheet['J13'] = data.get('cuadros_docente_instructor')

            worksheet['D17'] = data.get('total_tecnicos_fem')
            worksheet['E17'] = data.get('total_tecnicos_jovenes_t')    
            worksheet['F17'] = data.get('total_tecnicos_jovenes_t_fem')
            worksheet['G17'] = data.get('total_tecnicos_profe_titular')
            worksheet['H17'] = data.get('total_tecnicos_profe_aux')
            worksheet['I17'] = data.get('total_tecnicos_asistente')
            worksheet['J17'] =data.get('total_tecnicos_instructor')
            
            worksheet['D25'] = data.get('total_de_cuadros_fem')
            worksheet['E25'] = data.get('total_jovenes_t')      
            worksheet['F25'] = data.get('total_jovenes_t_fem') 
            worksheet['G25'] = data.get('profe_titular')
            worksheet['H25'] = data.get('profe_aux')
            worksheet['I25'] = data.get('asistente')
            worksheet['J25'] = data.get('instructor')

            #estilos
            bold_font = Font(bold=True, name='Areal',size=10)
            areal = Font(name = 'Areal', size=8)
            areal_b = Font(bold=True,name = 'Areal', size=8)
            center_alignment = Alignment(horizontal='center')
            worksheet.column_dimensions['A'].width = 30
            thick_border = Border(left=Side(style='medium'), 
                     right=Side(style='medium'), 
                     top=Side(style='medium'), 
                     bottom=Side(style='medium'))

            #estandarizaciones
            worksheet['A2'] = 'MODELO : 223.216(II) (AE3)'
            worksheet['A2'].font = bold_font
            worksheet['A4'] = 'INDICADORES TRIMESTRALES DE LA ACTIVIDAD DE RECURSOS HUMANOS'
            worksheet['A4'].font = bold_font
            worksheet['A6'] = 'CENTRO INFORMANTE :'
            worksheet['A6'].font = bold_font
            worksheet['A7'] = f'MES : {datetime.now().month} {datetime.now().year}'
            worksheet['A7'].font = bold_font
            worksheet['C8'] = 'PLANTILLA'
            worksheet['C8'].font = bold_font
            worksheet['A10'] = 'INDICADORES'
            worksheet['A10'].font = areal_b
            worksheet['A10'].alignment = center_alignment

            for row in worksheet['C8:Q8']:
               for cell in row:
                  cell.border = thick_border

            columns_to_adjust = {
              'B': 6,
              'C': 8,
              'D': 8,  
              'E': 6,  
              'F': 6,
              'G': 6,
              'H': 7,
              'I': 6,
              'J': 6,
              'K': 6,
              'L': 6,
              'M': 6,
              'N': 6,
              'O': 10,
              'P': 7,
              'Q': 6
            }

            for letter, width in columns_to_adjust.items():
              column = worksheet.column_dimensions[letter]
              column.width = width

            for row in worksheet['C9:Q9']:
               for cell in row:
                  cell.border = thick_border

            for row in worksheet['A10:Q26']:
               for cell in row:
                  cell.border = thick_border
            
            indc = {
               'A12' : '1. TOTAL CUADROS',
               'A13' : '2. Cuadros Docentes',
               'A14' : '3. Cuadros Administrativos',
               'A15' : '4. Cuadros Investigación',
               'A16' : '5. Otros Cuadros',
               'A17' : '6. TOTAL TÉCNICOS',
               'A18' : '7.  Profesores a Tiempo Completo',
               'A19' : '8. Asesores o Metodólogos',
               'A20' : '9.  Investigadores',
               'A21' : '10. Otros Técnicos',
               'A22' : '11. ADMINISTRATIVOS',
               'A23' : '12. SERVICIO',
               'A24' : '13. OPERARIOS',
               'A25' : 'TOTAL',
               'A26' : 'Profesores a Tiempo Parcial'
             }

            for key, value in indc.items():
              worksheet[key] = value
              worksheet[key].font = areal

            indc2 = {
               'C9' : 'Total',
               'D9' : 'De ellos:',
               'E9' : '   Jóvenes',
               'G9' : 'Profesores a Tiempo Completo e Investigadores'
            }

            for key, value in indc2.items():
              worksheet[key] = value
              worksheet[key].font = areal_b 

           
            for row in worksheet['A11:Q11']:
               for cell in row:
                  cell.value = str(self.count)
                  self.count += 1


            indc3 = {
               'B10': 'FILA',
               'C10': 'Cubierta',
               'D10': 'Fem',
               'E10': 'Total',
               'F10': 'Fem.',
               'G10': 'PT',
               'H10': 'PA',
               'I10': 'As',
               'J10': 'I',
               'K10': 'IT',
               'L10': 'IA',
               'M10': 'IAg',
               'N10': 'AI',
               'O10': 'Aux.Téc.Doc.',
               'P10': 'MsC',
               'Q10':'Dr.'
            }

            for key, value in indc3.items():
              worksheet[key] = value
              worksheet[key].font = areal_b 
             
            
            for row in worksheet['B12:B26']:
               for cell in row:
                  cell.value = str(self.count2)
                  self.count2 += 1

            for row in worksheet['C12:Q12']:
             for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            for row in worksheet['C17:Q17']:
             for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            for row in worksheet['C25:Q25']:
             for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

             for row in worksheet['C18:C20']:
              for cell in row:
                cell.fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            worksheet['C15'].fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            worksheet['C13'].fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            worksheet['C26'].fill = openpyxl.styles.PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

            worksheet['A28'] = 'Observaciones:'
            worksheet['A28'].font =bold_font
            worksheet['A31'] = f'Confeccionado por: {confe}'
            worksheet['A31'].font =bold_font
            worksheet['F31'] = f'Aprobado por: {aprob}'
            worksheet['F31'].font =bold_font

            filename = 'Ae3.xlsx'
            filepath = os.path.join(settings.BASE_DIR, 'static', 'informes', filename)
    
            workbook.save(filepath)

            
            history = History.objects.create( type_document = 'Ae3')
            history.save()
            
            ae3 = Ae3.objects.create(
                total_de_cuadros = data.get('total_de_cuadros'),
                cuadros_docente = data.get('cuadros_docente'),
                cuadros_admin = data.get('cuadros_admin'),
                cuadros_invest = data.get('cuadros_invest'),
                otros_cuadros = data.get('otros_cuadros'),
                total_tecnicos = data.get('total_tecnicos'),
                prof_tiempo_compl = data.get('prof_tiempo_compl'),
                asesores_metodol = data.get('asesores_metodol'),
                investigadores = data.get('investigadores'),
                otros_tecnicos = data.get('otros_tecnicos'),
                administrativos = data.get('administrativos'),
                servicio = data.get('servicio'),
                operarios = data.get('operarios'),
                total = data.get('total'),
                profe_tiempo_pars = data.get('profe_tiempo_pars'),
                total_de_cuadros_fem = data.get('total_de_cuadros_fem'),
                total_jovenes_t = data.get('total_jovenes_t'),
                total_jovenes_t_fem = data.get('total_jovenes_t_fem'),
                profe_titular = data.get('profe_titular'),
                profe_aux = data.get('profe_aux'),
                asistente = data.get('asistente'),
                instructor = data.get('instructor'),
                cuadros_docente_fem = data.get('cuadros_docente_fem'),
                cuadros_docente_jovenes_t = data.get('cuadros_docente_jovenes_t'),
                cuadros_docente_jovenes_t_fem = data.get('cuadros_docente_jovenes_t_fem'),
                cuadros_docente_profe_titular = data.get('cuadros_docente_profe_titular'),
                cuadros_docente_profe_aux = data.get('cuadros_docente_profe_aux'),
                cuadros_docente_asistente = data.get('cuadros_docente_asistente'),
                cuadros_docente_instructor = data.get('cuadros_docente_instructor'),
                total_tecnicos_fem = data.get('total_tecnicos_fem'),
                total_tecnicos_jovenes_t = data.get('total_tecnicos_jovenes_t'),
                total_tecnicos_jovenes_t_fem = data.get('total_tecnicos_jovenes_t_fem'),
                total_tecnicos_profe_titular = data.get('total_tecnicos_profe_titular'),
                total_tecnicos_profe_aux = data.get('total_tecnicos_profe_aux'),
                total_tecnicos_asistente = data.get('total_tecnicos_asistente'),
                total_tecnicos_instructor = data.get('total_tecnicos_instructor'),
                confeccionado = confe,
                aprobado = aprob,
            )
            ae3.save()
            
            return JsonResponse({
                'ok': True,
                'message': 'file created'
            })
        except Exception as error:
            return JsonResponse({
                'ok':False,
                'message': str(error)
            })